import openpyxl


class HomePageData:

    test_HomePage_data = [
        {"firstname": "Michał", "email": "eloelo321@o2.pl", "gender": "Male"},
        {"firstname": "Jim", "email": "bimbo@Pog.com", "gender": "Female"},
        {"firstname": "Kacper", "email": "Pog@gmail.com", "gender": "Male"},
    ]

    @staticmethod
    def getTestData(test_case_name): # will not use it because of  not excel
        Dict = {}
        book = openpyxl.load_workbook("Path")  # insert path
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=1, column=1).value == test_case_name:  # to make condition only get collums if its testcase 2
                for j in range(1, sheet.max_column + 1):  # to get colums
                    # Dict["Firstname] = "Kacper
                    Dict[sheet.cell(row=1, column=1).value] = sheet.cell(row=i, column=1).value

        return[Dict]


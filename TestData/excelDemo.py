import openpyxl

book = openpyxl.load_workbook("Path") # insert path / I dont have excel on my macbook
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Kacper"

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1): # to get rows
    if sheet.cell(row=1, column=1).value == "Testcase2": # to make condition only get collums if its testcase 2
        for j in range(1, sheet.max_column + 1): # to get colums
            # Dict["Firstname] = "Kacper
            Dict[sheet.cell(row=1, column=1).value] = sheet.cell(row=i, column=1).value